import asyncio
import logging
import sys
import os
import json
import pandas as pd
import time
import requests
from datetime import datetime, timedelta, date
from zoneinfo import ZoneInfo
from typing import Any, Dict, List, Optional, Set, Tuple
import traceback
from openpyxl import load_workbook
from market.tick_subscriber import TickSubscriber
from market.option_engine import OptionEngine

# === CONSTANTS ===
IST_ZONE = "Asia/Kolkata"
CSV_FLOAT_FORMAT = "%.2f"
OPTION_PREMIUM_ACCEPT = (50.0, 150.0)
NUM_CLOSEST_STRIKES = 3
MAX_ERRORS = 5
STRIKE_DIVISOR = 100

# === FILE PATHS ===
# SCRIPT_DIR = "/home/algo/strategies/ulpl"
SCRIPT_DIR = r"E:\SHARMKT\ALGO\strategies\ulpl"
LOG_FILE = os.path.join(SCRIPT_DIR, "ulpl.log")
CREDENTIALS_PATH = os.path.join(SCRIPT_DIR, "config_ulpl.json")
SYMBOL_PATH = os.path.join(SCRIPT_DIR, "symbol_ulpl.json")

# KEYDOCS_DIR = "/home/ubuntu/keydocs"
KEYDOCS_DIR = r"E:\SHARMKT\PYTHON STARTGY\keydocs"
HOLIDAYS_PATH = os.path.join(KEYDOCS_DIR, "holidays.json")

MATCHED_CSV_PATH: str = ""

INDICATORS_DIR = r"E:\SHARMKT\ALGO\data\data\indicators"
NIFTY_INDICATORS_PATH = os.path.join(INDICATORS_DIR, "NIFTY.json")

TRADE_LOG_PATH = os.path.join(SCRIPT_DIR, "Paper_ulpl.xlsx")
TRADE_LOG_COLUMNS = [
    "symbol", "token", "open", "close", "trend", "qty", "entry date&time",
    "sell", "buy", "exit date&time", "p&l", "entry reason", "exit reason"
]
TRADE_LOG_COLUMNS_DTYPES = {
    "symbol": "object",
    "token": "object",
    "open": "float64",
    "close": "float64",
    "trend": "object",
    "qty": "int64",
    "entry date&time": "object",
    "sell": "float64",
    "buy": "float64",
    "exit date&time": "object",
    "p&l": "float64",
    "entry reason": "object",
    "exit reason": "object",
}

# === LOGGING ===
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

if sys.platform.startswith("win"):
    import _locale
    _locale._getdefaultlocale = (lambda *args: ["en_US", "utf8"])


# === UTILITY FUNCTIONS ===
def send_telegram_message(bot_token: str, chat_id: str, message: str) -> None:
    try:
        url = f"https://api.telegram.org/bot{bot_token}/sendMessage"
        payload = {"chat_id": chat_id, "text": message, "parse_mode": "HTML"}
        response = requests.post(url, json=payload, timeout=10)
        if response.status_code == 200:
            logger.info(f"Telegram message sent: {message}")
        else:
            logger.error(f"Telegram error: {response.text}")
    except Exception as e:
        logger.error(f"Error sending Telegram message: {e}\n{traceback.format_exc()}")


def load_holidays() -> Set[date]:
    try:
        with open(HOLIDAYS_PATH, "r") as f:
            holidays_data = json.load(f)
        return set(datetime.strptime(d, "%Y-%m-%d").date() for d in holidays_data["holidays"])
    except Exception as e:
        logger.error(f"Failed to load holidays: {e}\n{traceback.format_exc()}")
        return set()


def is_holiday(d: date, holidays: Set[date], weekly_off_days: Optional[List[str]] = None) -> bool:
    weekly_off_days = weekly_off_days or ["SAT", "SUN"]
    day_name = d.strftime("%a").upper()[:3]
    return day_name in weekly_off_days or d in holidays


def is_trading_day(today: date, holidays: Set[date], weekly_off_days: Optional[List[str]] = None) -> bool:
    return not is_holiday(today, holidays, weekly_off_days)


def load_configurations() -> Tuple[
    str, int, str, str, Dict[str, Any], bool, int, Dict[str, Any], Dict[str, Any], str
]:
    try:
        with open(CREDENTIALS_PATH, "r") as f:
            config = json.load(f)

        with open(SYMBOL_PATH, "r") as f:
            symbol_config = json.load(f)

        required_keys = ["telegram_bot_token", "telegram_chat_id", "matched_csv_path"]
        for key in required_keys:
            if key not in config:
                raise KeyError(f"Missing key in credentials config: {key}")

        sma_period = config.get("SMA_PERIOD", 12)
        interval = config.get("interval", "THIRTY_MINUTE")
        paper_trading = config.get("paper_trading", True)
        paper_lot = config.get("paper_lot", 1)
        market = config.get("market", {})
        intervals = config.get("intervals", {})
        matched_csv_path = config["matched_csv_path"]

        return (
            interval,
            sma_period,
            config["telegram_bot_token"],
            config["telegram_chat_id"],
            symbol_config,
            paper_trading,
            paper_lot,
            market,
            intervals,
            matched_csv_path
        )
    except Exception as e:
        logger.error(f"Failed to load configurations: {e}\n{traceback.format_exc()}")
        raise


def get_interval_minutes(intervals_config: Dict[str, Any], interval: str) -> int:
    return int(intervals_config.get(interval, {}).get("minutes", 30))


def get_interval_key(intervals_config: Dict[str, Any], interval: str) -> str:
    return intervals_config.get(interval, {}).get("key", "30min")


def get_market_time(market_config: Dict[str, Any], key: str, default_value: str) -> str:
    return market_config.get(key, default_value)


def get_last_closed_candle_time(
    current_time: datetime,
    interval: str,
    intervals_config: Dict[str, Any],
    market_config: Dict[str, Any]
) -> datetime:
    interval_minutes = get_interval_minutes(intervals_config, interval)
    start_time = get_market_time(market_config, "start_time", "00:15")
    market_open = datetime.combine(
        current_time.date(),
        datetime.strptime(start_time, "%H:%M").time(),
        tzinfo=current_time.tzinfo
    )
    minutes_since_open = (current_time - market_open).total_seconds() / 60
    if minutes_since_open < 0:
        return market_open
    completed_intervals = int(minutes_since_open // interval_minutes)
    last_closed = market_open + timedelta(minutes=completed_intervals * interval_minutes)
    return last_closed


def get_next_candle_close_time(
    current_time: datetime,
    interval: str,
    intervals_config: Dict[str, Any],
    market_config: Dict[str, Any]
) -> datetime:
    interval_minutes = get_interval_minutes(intervals_config, interval)
    last_closed = get_last_closed_candle_time(current_time, interval, intervals_config, market_config)
    return last_closed + timedelta(minutes=interval_minutes)


def load_nifty_indicators(interval: str, intervals_config: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    try:
        if not os.path.exists(NIFTY_INDICATORS_PATH):
            logger.error(f"Indicators file not found: {NIFTY_INDICATORS_PATH}")
            return None

        with open(NIFTY_INDICATORS_PATH, "r", encoding="utf-8") as f:
            data = json.load(f)

        tf = get_interval_key(intervals_config, interval)
        if not tf:
            logger.error(f"Unsupported interval for indicators JSON: {interval}")
            return None

        indicators = data.get("indicators", {})
        tf_data = indicators.get(tf)
        if not tf_data:
            logger.error(f"No indicator data found for timeframe: {tf}")
            return None

        candle = tf_data.get("candle", {})
        close_price = candle.get("close")
        sma_value = tf_data.get("sma12")

        if close_price is None:
            logger.error(f"Close price missing in indicators file for {tf}")
            return None

        candle_ts = pd.to_datetime(candle.get("timestamp"))
        if candle_ts.tzinfo is None:
            candle_ts = candle_ts.tz_localize(ZoneInfo(IST_ZONE))
        else:
            candle_ts = candle_ts.tz_convert(ZoneInfo(IST_ZONE))

        trend = None
        if sma_value is not None:
            if close_price > sma_value:
                trend = "bullish"
            elif close_price < sma_value:
                trend = "bearish"

        return {
            "timeframe": tf,
            "updated_at": data.get("updated_at"),
            "last_candle": data.get("last_candle"),
            "timestamp": candle_ts,
            "open": candle.get("open"),
            "high": candle.get("high"),
            "low": candle.get("low"),
            "close": close_price,
            "volume": candle.get("volume", 0),
            "sma": sma_value,
            "trend": trend,
            "raw": tf_data
        }

    except Exception as e:
        logger.error(f"Failed to load NIFTY indicators: {e}\n{traceback.format_exc()}")
        return None


def get_current_nifty_expiry(
    matched_csv_path: str,
    today: date,
    holidays: Set[date],
    weekly_off_days: Optional[List[str]] = None
):
    if not os.path.exists(matched_csv_path):
        raise FileNotFoundError(f"matched.csv not found: {matched_csv_path}")

    df = pd.read_csv(matched_csv_path)

    nifty_opts = df[
        (df["a_name"] == "NIFTY") &
        (df["a_exch_seg"] == "NFO") &
        (df["a_instrumenttype"] == "OPTIDX")
    ].copy()

    nifty_opts["expiry_date"] = pd.to_datetime(
        nifty_opts["a_expiry"], format="%d%b%Y", errors="coerce"
    )
    nifty_opts.dropna(subset=["expiry_date"], inplace=True)
    expiry_list = sorted(nifty_opts["expiry_date"].dt.date.unique())
    future_expiries = [e for e in expiry_list if e >= today]

    if not future_expiries:
        raise ValueError("No future NIFTY expiries found in matched.csv")

    nearest_expiry = future_expiries[0]

    def future_trading_days(start: date, end: date) -> int:
        days = 0
        d = start + timedelta(days=1)
        while d <= end:
            if not is_holiday(d, holidays, weekly_off_days):
                days += 1
            d += timedelta(days=1)
        return days

    remaining_sessions = future_trading_days(today, nearest_expiry)
    if remaining_sessions <= 1 and len(future_expiries) > 1:
        nearest_expiry = future_expiries[1]

    lot_row = nifty_opts[nifty_opts["expiry_date"].dt.date == nearest_expiry].iloc[0]
    lot_size = int(lot_row["a_lotsize"])

    expiry_str = nearest_expiry.strftime("%d%b%y").upper()
    return expiry_str, lot_size


def load_instrument_tokens(matched_csv_path: str, expiry_str: str) -> List[Dict[str, Any]]:
    try:
        if not os.path.exists(matched_csv_path):
            raise FileNotFoundError(f"matched.csv not found: {matched_csv_path}")

        df = pd.read_csv(matched_csv_path)

        expiry_date = datetime.strptime(expiry_str, "%d%b%y").date()
        expiry_full = expiry_date.strftime("%d%b%Y").upper()

        nifty_options = df[
            (df["a_exch_seg"] == "NFO") &
            (df["a_name"] == "NIFTY") &
            (df["a_instrumenttype"] == "OPTIDX") &
            (df["a_expiry"].str.upper() == expiry_full) &
            (df["a_symbol"].str.endswith(("CE", "PE"), na=False))
        ].copy()

        nifty_options["token"] = pd.to_numeric(nifty_options["token"], errors="coerce")
        nifty_options["strike"] = pd.to_numeric(nifty_options["a_strike"], errors="coerce") / STRIKE_DIVISOR
        nifty_options.dropna(subset=["token", "strike"], inplace=True)
        nifty_options = nifty_options.rename(columns={"a_symbol": "symbol"})

        logger.info(f"Loaded {len(nifty_options)} NIFTY options for expiry {expiry_str} from matched.csv")
        return nifty_options[["symbol", "token", "strike"]].to_dict("records")

    except Exception as e:
        logger.error(f"Error loading instruments from matched.csv: {e}\n{traceback.format_exc()}")
        return []


def filter_strikes(current_price: float, option_data: List[Dict[str, Any]], single_type: Optional[str] = None):
    if not option_data:
        return (None, None) if single_type is None else None

    ces = [o for o in option_data if o["type"] == "CE"]
    pes = [o for o in option_data if o["type"] == "PE"]

    premium_ranges = [(80.0, 100.0), (70.0, 100.0), (70.0, 110.0)]
    round_prefs = [100, 50]

    def find_best(options: List[Dict[str, Any]], price: float) -> Optional[Dict[str, Any]]:
        if not options:
            return None
        for min_p, max_p in premium_ranges:
            candidates = [o for o in options if min_p <= o["premium"] <= max_p]
            if not candidates:
                continue
            for mod in round_prefs:
                round_cand = [o for o in candidates if o["strike"] % mod == 0]
                if round_cand:
                    return min(round_cand, key=lambda o: abs(o["strike"] - price))
            return min(candidates, key=lambda o: abs(o["strike"] - price))
        return None

    if single_type == "CE":
        return find_best(ces, current_price)
    if single_type == "PE":
        return find_best(pes, current_price)

    ce = find_best(ces, current_price)
    pe = find_best(pes, current_price)
    return ce, pe


def load_trade_log() -> pd.DataFrame:
    if os.path.exists(TRADE_LOG_PATH):
        try:
            df = pd.read_excel(TRADE_LOG_PATH)
            for col, dtype in TRADE_LOG_COLUMNS_DTYPES.items():
                if col in df and df[col].dtype != dtype:
                    df[col] = df[col].astype(dtype, errors="ignore")
            return df
        except Exception as e:
            logger.error(f"Failed to read trade log: {e}")

    df = pd.DataFrame({col: pd.Series(dtype=dtype) for col, dtype in TRADE_LOG_COLUMNS_DTYPES.items()})
    df.to_excel(TRADE_LOG_PATH, index=False)
    return df


def save_trade_log(df: pd.DataFrame):
    df.to_excel(TRADE_LOG_PATH, index=False)
    col_widths = {
        "symbol": 21,
        "entry date&time": 17.57,
        "exit date&time": 17.57,
        "entry reason": 14,
        "exit reason": 14,
        "open": 10,
        "close": 10,
        "SL": 10
    }

    wb = load_workbook(TRADE_LOG_PATH)
    ws = wb.active
    for col in ws.iter_cols(1, ws.max_column):
        header = col[0].value
        if header in col_widths:
            ws.column_dimensions[col[0].column_letter].width = col_widths[header]
    wb.save(TRADE_LOG_PATH)


def add_trade_log_row(trade: Dict[str, Any]):
    df = load_trade_log()
    row_df = pd.DataFrame([trade])
    for col, dtype in TRADE_LOG_COLUMNS_DTYPES.items():
        if col in row_df and row_df[col].dtype != dtype:
            row_df[col] = row_df[col].astype(dtype, errors="ignore")
    if df.empty:
        df = row_df
    else:
        df = pd.concat([df, row_df], ignore_index=True)
    save_trade_log(df)


def update_trade_exit(symbol: str, token: str, exit_time: str, exit_price: float, p_and_l: float, exit_reason: str):
    df = load_trade_log()
    symbol = str(symbol)
    token = str(token)

    exit_dt_col = df["exit date&time"].astype(str)
    is_exit_null = exit_dt_col.isnull() | (exit_dt_col == "") | (exit_dt_col.str.lower() == "nan")

    cond = (
        (df["symbol"].astype(str) == symbol) &
        (df["token"].astype(str) == token) &
        is_exit_null
    )

    idx = df[cond].index
    if not idx.empty:
        i = idx[0]
        df.at[i, "exit date&time"] = str(exit_time)
        df.at[i, "buy"] = float(exit_price)
        df.at[i, "p&l"] = float(p_and_l)
        df.at[i, "exit reason"] = str(exit_reason)
        save_trade_log(df)
        logger.info("Trade exit updated and saved.")
    else:
        logger.warning("No matching open trade found to update exit.")


def get_active_trades() -> pd.DataFrame:
    df = load_trade_log()
    return df[df["exit date&time"].isnull()]


def log_trade_entry(
    symbol: str,
    token: str,
    open_price: float,
    close_price: float,
    trend: str,
    qty: int,
    entry_time: str,
    sell_price: float,
    entry_reason: str
):
    trade = {
        "symbol": symbol,
        "token": token,
        "open": open_price,
        "close": close_price,
        "trend": trend,
        "qty": qty,
        "entry date&time": entry_time,
        "sell": sell_price,
        "buy": None,
        "exit date&time": None,
        "p&l": None,
        "entry reason": entry_reason,
        "exit reason": None
    }
    add_trade_log_row(trade)


class PaperTrader:
    def __init__(
        self,
        interval: str,
        sma_period: int,
        symbol_config: Dict[str, Any],
        telegram_bot_token: str,
        telegram_chat_id: str,
        paper_trading: bool,
        paper_lot: int,
        market_config: Dict[str, Any],
        intervals_config: Dict[str, Any],
        matched_csv_path: str
    ):
        self.interval = interval
        self.sma_period = sma_period
        self.symbol_config = symbol_config
        self.telegram_bot_token = telegram_bot_token
        self.telegram_chat_id = telegram_chat_id
        self.paper_trading = paper_trading
        self.paper_lot = paper_lot
        self.market_config = market_config or {}
        self.intervals_config = intervals_config or {}
        self.matched_csv_path = matched_csv_path

        self.last_pnl_report_time = None
        self.active_trades: Dict[str, Dict[str, Any]] = {}
        self.last_trend = None
        self.last_close_price = None
        self.last_entry_time = None

        self.lot_size = 0
        self.qty = 0

        self.last_entry_trend = None
        self.time_exit_done = False
        self.wait_for_new_candle = False
        self.current_prices: Dict[str, float] = {}
        self.both_sl_hit_in_trend = False
        self.entry_open = None
        self.entry_close = None
        self.logger = logging.getLogger(__name__)
        self.expiry_str = None
        self.sl_monitor = None
        self.option_engine = None

    def send_msg(self, msg: str):
        send_telegram_message(self.telegram_bot_token, self.telegram_chat_id, msg)

    def get_active_trades_dict(self):
        return self.active_trades.copy()

    def update_ltp(self, token: str, ltp: float):
        token = str(token)
        self.current_prices[token] = ltp
        trade = self.active_trades.get(token)

        if token in self.active_trades:
            print("LTP update:", token, ltp)

        if trade:
            sl_price = trade["sell"] * 1.20
            if ltp >= sl_price:
                self.handle_sl_exit(trade, ltp)

    def get_current_ltp(self, token: str, default: float) -> float:
        return self.current_prices.get(str(token), default)

    def handle_sl_exit(self, trade: Dict[str, Any], ltp: float):
        token_str = str(trade["token"])

        if token_str not in self.active_trades:
            return

        p_and_l = (trade["sell"] - ltp) * trade["qty"]

        update_trade_exit(
            trade["symbol"],
            trade["token"],
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            ltp,
            p_and_l,
            "SL Hit"
        )

        self.send_msg(
            f"SL Exit: {trade['symbol']} @ {ltp:.2f}, "
            f"SL={trade['sell'] * 1.20:.2f}, "
            f"P&L={p_and_l:.2f}"
        )

        del self.active_trades[token_str]

        if not self.active_trades:
            self.both_sl_hit_in_trend = True
            self.logger.info(f"Both legs SL hit for trend {self.last_trend}")

    def wait_for_initial_ltp(self, timeout: int = 10) -> bool:
        tokens = set(self.active_trades.keys())
        start = time.time()
        while time.time() - start < timeout:
            missing = [t for t in tokens if t not in self.current_prices]
            if not missing:
                return True
            time.sleep(0.5)
        self.logger.warning(f"Timeout waiting for LTPs for tokens: {missing}")
        return False

    def calculate_running_pnl(self) -> str:
        total_pnl = 0.0
        details = []
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        for token, trade in self.active_trades.items():
            ltp = self.get_current_ltp(token, trade["sell"])
            pnl = (trade["sell"] - ltp) * trade["qty"]
            total_pnl += pnl
            sl_price = trade["sell"] * 1.20

            details.append(
                f"{trade['symbol']} | {trade['qty']} | {trade['sell']:.2f} | {ltp:.2f} | "
                f"SL: {sl_price:.2f} | P&L: {pnl:+.2f}"
            )

        msg = f"P&L Report @ {now}\n\n"
        if details:
            msg += "\n".join(details)
            msg += f"\n\nTotal P&L: {total_pnl:+.2f}"
        else:
            msg += "No active trades."

        return msg

    async def _enter_option_trade(
        self,
        now: datetime,
        candle: pd.Series,
        trend: str,
        entry_reason: str,
        single_type: Optional[str] = None,
        prefix: str = "New paper trade"
    ) -> bool:
        try:
            if not self.paper_trading:
                self.logger.info("paper_trading is disabled. Skipping paper trade entry.")
                return False

            close_price = float(candle["close"])

            option_data = self.option_engine.get_range(80, 100)
            if not option_data:
                option_data = self.option_engine.get_range(70, 100)
            if not option_data:
                option_data = self.option_engine.get_range(70, 110)
            if not option_data:
                self.logger.warning("No options found in premium range")
                return False

            entry_time = now.strftime("%Y-%m-%d %H:%M:%S")
            open_price = float(candle["open"])
            entered = False

            if single_type:
                selected = filter_strikes(close_price, option_data, single_type=single_type)
                if selected is not None:
                    log_trade_entry(
                        selected["symbol"],
                        selected["token"],
                        open_price,
                        close_price,
                        trend,
                        self.qty,
                        entry_time,
                        selected["premium"],
                        entry_reason
                    )
                    self.active_trades[str(selected["token"])] = {
                        "symbol": selected["symbol"],
                        "token": selected["token"],
                        "sell": selected["premium"],
                        "qty": self.qty,
                        "entry_time": entry_time
                    }
                    self.send_msg(
                        f"{prefix}: {single_type} {selected['symbol']}@{selected['premium']:.2f}, "
                        f"Trend={trend}, Entry Reason={entry_reason}"
                    )
                    entered = True
                else:
                    self.logger.warning(f"No suitable {single_type} found for {entry_reason}")
            else:
                ce_option, pe_option = filter_strikes(close_price, option_data)
                ce_str = pe_str = "-"
                ce_prem = pe_prem = "-"

                if ce_option is not None:
                    log_trade_entry(
                        ce_option["symbol"],
                        ce_option["token"],
                        open_price,
                        close_price,
                        trend,
                        self.qty,
                        entry_time,
                        ce_option["premium"],
                        entry_reason
                    )
                    self.active_trades[str(ce_option["token"])] = {
                        "symbol": ce_option["symbol"],
                        "token": ce_option["token"],
                        "sell": ce_option["premium"],
                        "qty": self.qty,
                        "entry_time": entry_time
                    }
                    ce_str = ce_option["symbol"]
                    ce_prem = f"{ce_option['premium']:.2f}"
                    entered = True

                if pe_option is not None:
                    log_trade_entry(
                        pe_option["symbol"],
                        pe_option["token"],
                        open_price,
                        close_price,
                        trend,
                        self.qty,
                        entry_time,
                        pe_option["premium"],
                        entry_reason
                    )
                    self.active_trades[str(pe_option["token"])] = {
                        "symbol": pe_option["symbol"],
                        "token": pe_option["token"],
                        "sell": pe_option["premium"],
                        "qty": self.qty,
                        "entry_time": entry_time
                    }
                    pe_str = pe_option["symbol"]
                    pe_prem = f"{pe_option['premium']:.2f}"
                    entered = True

                self.send_msg(
                    f"{prefix}: CE {ce_str}@{ce_prem}, PE {pe_str}@{pe_prem}, "
                    f"Trend={trend}, Entry Reason={entry_reason}"
                )

            if entered:
                self.last_entry_trend = trend
                self.entry_open = open_price
                self.entry_close = close_price

                if self.active_trades and self.sl_monitor is not None:
                    self.sl_monitor.start()
                    self.logger.info(f"SLWebSocketMonitor restarted after {entry_reason}")

                self.wait_for_initial_ltp(timeout=10)
                return True

            return False

        except Exception as e:
            self.logger.error(f"Error in _enter_option_trade ({entry_reason}): {e}")
            self.send_msg(f"❌ Entry error ({entry_reason}): {str(e)}")
            traceback.print_exc()
            return False

    async def run(self, ist: ZoneInfo, holidays: Set[date], today: date, sl_monitor):
        self.sl_monitor = sl_monitor

        market_enabled = self.market_config.get("enabled", True)
        weekly_off_days = self.market_config.get("weekly_off_days", ["SAT", "SUN"])

        start_time = get_market_time(self.market_config, "start_time", "00:15")
        end_time = get_market_time(self.market_config, "end_time", "15:30")
        exit_time_str = get_market_time(self.market_config, "exit_time", "15:15")

        market_open_time = datetime.combine(
            today, datetime.strptime(start_time, "%H:%M").time(), tzinfo=ist
        )
        market_close_time = datetime.combine(
            today, datetime.strptime(end_time, "%H:%M").time(), tzinfo=ist
        )
        exit_time = datetime.combine(
            today, datetime.strptime(exit_time_str, "%H:%M").time(), tzinfo=ist
        )

        interval_key = get_interval_key(self.intervals_config, self.interval)
        interval_minutes = get_interval_minutes(self.intervals_config, self.interval)

        self.logger.info(f"Using interval={self.interval}, key={interval_key}, minutes={interval_minutes}")

        expiry_str, lot_size = get_current_nifty_expiry(
            self.matched_csv_path, today, holidays, weekly_off_days
        )

        self.expiry_str = expiry_str
        self.lot_size = lot_size
        self.qty = self.paper_lot * self.lot_size

        instruments = load_instrument_tokens(self.matched_csv_path, self.expiry_str)
        self.option_engine = OptionEngine(instruments)

        self.logger.info(f"Paper trading enabled: {self.paper_trading}")
        self.logger.info(f"Paper lot: {self.paper_lot}, Final Qty per leg: {self.qty}")
        self.logger.info(f"Using Expiry: {expiry_str}, Lot Size: {lot_size}")
        self.send_msg(f"Trading Expiry: {expiry_str} | Lot Size: {lot_size}")

        nifty_symbol = next((s for s in self.symbol_config if "NIFTY" in s.upper()), None)

        df_active = get_active_trades()
        for _, row in df_active.iterrows():
            self.active_trades[str(row["token"])] = dict(row)

        if not df_active.empty:
            self.entry_open = df_active["open"].iloc[0]
            self.entry_close = df_active["close"].iloc[0]
            self.last_entry_trend = df_active["trend"].iloc[0]
            self.logger.info(f"Resuming with {len(self.active_trades)} active trades")

        if not self.active_trades:
            self.wait_for_new_candle = True

        self.wait_for_initial_ltp(timeout=10)

        last_processed_candle = None
        no_new_ohlc_counter = 0

        while True:
            now = datetime.now(ist)

            if not market_enabled:
                self.logger.info("Market config disabled. Exiting run loop.")
                break

            if now < market_open_time:
                self.logger.info(f"Waiting until Market open @ {market_open_time}")
                time.sleep((market_open_time - now).total_seconds())
                continue

            if now >= market_close_time:
                self.send_msg("Market closed for the day, Exiting.")
                break

            last_candle_time = get_last_closed_candle_time(
                now, self.interval, self.intervals_config, self.market_config
            )
            target_candle_time = last_candle_time - timedelta(minutes=interval_minutes)
            next_candle_time = get_next_candle_close_time(
                now, self.interval, self.intervals_config, self.market_config
            )

            if (
                target_candle_time != last_processed_candle and
                (now >= next_candle_time or (next_candle_time - now).total_seconds() > -30)
            ):
                self.logger.info(f"Processing data for candle closed at {target_candle_time}")
                last_processed_candle = target_candle_time
                retry_for_this_candle = 0
                indicator_data = None

                while retry_for_this_candle < 5:
                    indicator_data = load_nifty_indicators(self.interval, self.intervals_config)
                    if indicator_data is not None:
                        no_new_ohlc_counter = 0
                        break

                    retry_for_this_candle += 1
                    if retry_for_this_candle == 5:
                        no_new_ohlc_counter += 1
                        self.send_msg(
                            f"Failed to load indicator data for {nifty_symbol} "
                            f"after 5 attempts at {target_candle_time}."
                        )
                        if no_new_ohlc_counter == 2:
                            self.send_msg(
                                f"Exiting script: Failed to load indicator data for "
                                f"two consecutive candles at {target_candle_time}."
                            )
                            return
                        sleep_seconds = (next_candle_time - datetime.now(ist)).total_seconds()
                        self.logger.info(f"Sleeping until next candle: {sleep_seconds:.2f} seconds.")
                        if sleep_seconds > 0:
                            time.sleep(sleep_seconds)
                        break
                    else:
                        self.logger.warning(
                            f"No indicator data for {nifty_symbol}. Retry {retry_for_this_candle}/5."
                        )
                        time.sleep(10)

                if indicator_data is None:
                    continue

                candle_ts = pd.to_datetime(indicator_data["timestamp"])
                if candle_ts.tzinfo is None:
                    candle_ts = candle_ts.tz_localize(ist)
                else:
                    candle_ts = candle_ts.tz_convert(ist)

                candle = pd.Series({
                    "timestamp": candle_ts,
                    "open": float(indicator_data["open"]),
                    "high": float(indicator_data["high"]),
                    "low": float(indicator_data["low"]),
                    "close": float(indicator_data["close"]),
                    "volume": float(indicator_data["volume"] or 0),
                })

                close_price = candle["close"]
                trend = indicator_data["trend"]

                previous_trend = self.last_trend
                trend_change = (
                    previous_trend is not None and
                    trend is not None and
                    previous_trend != trend
                )
                self.last_trend = trend
                candle_time = candle["timestamp"]
                self.last_close_price = close_price

                self.logger.info(
                    f"Trend change: {trend_change}, Trend: {trend}, "
                    f"SMA: {indicator_data.get('sma')}, Candle Time: {indicator_data.get('timestamp')}"
                )

                if now >= exit_time and self.active_trades:
                    for token, trade in list(self.active_trades.items()):
                        ltp = self.get_current_ltp(token, trade["sell"])
                        p_and_l = (trade["sell"] - ltp) * trade["qty"]
                        update_trade_exit(
                            trade["symbol"],
                            trade["token"],
                            now.strftime("%Y-%m-%d %H:%M:%S"),
                            ltp,
                            p_and_l,
                            "Time Exit"
                        )
                        self.send_msg(f"Time exit: {trade['symbol']} @ {ltp:.2f}, P&L={p_and_l:.2f}")
                    self.active_trades = {}
                    self.time_exit_done = True
                    self.logger.info("Time-based exit done. Continuing to process candles until market close.")
                    break

                if now >= exit_time:
                    self.logger.info(f"No new trades or data fetching after market exit time ({exit_time_str}).")
                    sleep_seconds = (market_close_time - now).total_seconds()
                    self.logger.info(
                        f"Sleeping until market close at {market_close_time} for {sleep_seconds:.2f} seconds."
                    )
                    if sleep_seconds > 0:
                        time.sleep(sleep_seconds)
                    break

                if trend_change:
                    if not self.active_trades:
                        if self.wait_for_new_candle:
                            next_candle_time = get_next_candle_close_time(
                                now, self.interval, self.intervals_config, self.market_config
                            )
                            sleep_seconds = (next_candle_time - now).total_seconds()
                            if sleep_seconds > 1:
                                self.logger.info(
                                    f"Waiting {sleep_seconds:.2f} sec for next candle close before entering new trades"
                                )
                                time.sleep(sleep_seconds)
                            self.wait_for_new_candle = False
                            continue

                        if (now - candle_time).total_seconds() < 0:
                            time.sleep(5)
                            continue

                        await self._enter_option_trade(
                            now, candle, trend, "Trend Change", prefix="New paper trade"
                        )

                    else:
                        ce_ltp = pe_ltp = None
                        ce_trade = pe_trade = None

                        for token, trade in self.active_trades.items():
                            if trade["symbol"].endswith("CE"):
                                ce_ltp = self.get_current_ltp(token, trade["sell"])
                                ce_trade = trade
                            elif trade["symbol"].endswith("PE"):
                                pe_ltp = self.get_current_ltp(token, trade["sell"])
                                pe_trade = trade

                        active_trade_count = len(self.active_trades)
                        ce_in_range = ce_ltp is not None and 80.0 <= ce_ltp <= 100.0
                        pe_in_range = pe_ltp is not None and 80.0 <= pe_ltp <= 100.0

                        if active_trade_count == 1:
                            trade = ce_trade if ce_trade else pe_trade
                            ltp = ce_ltp if ce_trade else pe_ltp
                            in_range = ce_in_range if ce_trade else pe_in_range
                            missing_leg = "PE" if ce_trade else "CE"

                            if in_range:
                                self.logger.info(
                                    f"Single trade {trade['symbol']} LTP {ltp:.2f} "
                                    f"in 80-100 range during trend reversal, continuing."
                                )
                                self.send_msg(
                                    f"Continuing {trade['symbol']} @ {ltp:.2f} "
                                    f"due to trend reversal (LTP in 80-100 range)."
                                )
                                await self._enter_option_trade(
                                    now,
                                    candle,
                                    trend,
                                    "80-100 Trend Reversal",
                                    single_type=missing_leg,
                                    prefix="Re-entry paper trade"
                                )
                                self.both_sl_hit_in_trend = False
                            else:
                                ltp = self.get_current_ltp(trade["token"], trade["sell"])
                                p_and_l = (trade["sell"] - ltp) * trade["qty"]
                                update_trade_exit(
                                    trade["symbol"],
                                    trade["token"],
                                    now.strftime("%Y-%m-%d %H:%M:%S"),
                                    ltp,
                                    p_and_l,
                                    "Trend Reversal - LTP Out of Range"
                                )
                                self.send_msg(
                                    f"Trend reversal exit: {trade['symbol']} @ {ltp:.2f}, P&L={p_and_l:.2f}"
                                )
                                del self.active_trades[str(trade["token"])]
                                await self._enter_option_trade(
                                    now,
                                    candle,
                                    trend,
                                    "New Trend",
                                    prefix="New paper trade after reversal"
                                )

                        elif active_trade_count == 2:
                            self.logger.info(f"Trend reversal premiums → CE={ce_ltp:.2f} | PE={pe_ltp:.2f}")
                            self.send_msg(f"Trend reversal premiums → CE={ce_ltp:.2f} | PE={pe_ltp:.2f}")

                            if not ce_in_range and not pe_in_range:
                                for token, trade in list(self.active_trades.items()):
                                    ltp = self.get_current_ltp(token, trade["sell"])
                                    p_and_l = (trade["sell"] - ltp) * trade["qty"]
                                    update_trade_exit(
                                        trade["symbol"],
                                        trade["token"],
                                        now.strftime("%Y-%m-%d %H:%M:%S"),
                                        ltp,
                                        p_and_l,
                                        "Trend Reversal - LTP Out of Range"
                                    )
                                    self.send_msg(
                                        f"Trend reversal exit: {trade['symbol']} @ {ltp:.2f}, P&L={p_and_l:.2f}"
                                    )
                                    del self.active_trades[token]

                                await self._enter_option_trade(
                                    now,
                                    candle,
                                    trend,
                                    "New Trend",
                                    prefix="New paper trade after reversal"
                                )

                            elif ce_in_range != pe_in_range:
                                exit_trade = pe_trade if ce_in_range else ce_trade
                                missing_leg = "PE" if ce_in_range else "CE"
                                ltp = self.get_current_ltp(exit_trade["token"], exit_trade["sell"])
                                p_and_l = (exit_trade["sell"] - ltp) * exit_trade["qty"]
                                update_trade_exit(
                                    exit_trade["symbol"],
                                    exit_trade["token"],
                                    now.strftime("%Y-%m-%d %H:%M:%S"),
                                    ltp,
                                    p_and_l,
                                    "Trend Reversal - LTP Out of Range"
                                )
                                self.send_msg(
                                    f"Trend reversal exit: {exit_trade['symbol']} @ {ltp:.2f}, P&L={p_and_l:.2f}"
                                )
                                del self.active_trades[str(exit_trade["token"])]
                                await self._enter_option_trade(
                                    now,
                                    candle,
                                    trend,
                                    "Trend Reversal",
                                    single_type=missing_leg,
                                    prefix="Re-entry paper trade"
                                )
                            else:
                                self.logger.info(
                                    f"Both trades in 80-100 range (CE: {ce_ltp:.2f}, PE: {pe_ltp:.2f}), continuing."
                                )
                                self.both_sl_hit_in_trend = False
                        else:
                            self.logger.warning(f"Unexpected number of active trades: {active_trade_count}")
                            self.both_sl_hit_in_trend = False

                else:
                    if self.entry_open is not None and self.entry_close is not None:
                        lower_price = min(self.entry_open, self.entry_close)
                        high_price = max(self.entry_open, self.entry_close)
                        ce_re_enter = close_price < lower_price
                        pe_re_enter = close_price > high_price

                        if len(self.active_trades) == 1:
                            remaining_token = list(self.active_trades.keys())[0]
                            remaining_symbol = self.active_trades[remaining_token]["symbol"]
                            missing_leg = "PE" if remaining_symbol.endswith("CE") else "CE"
                            re_enter_status = ce_re_enter if missing_leg == "CE" else pe_re_enter

                            self.logger.info(
                                f"Re-entry check for {missing_leg}: Condition {'met' if re_enter_status else 'not met'}, "
                                f"Entry Open={self.entry_open:.2f}, Entry Close={self.entry_close:.2f}, "
                                f"Current Close={close_price:.2f}"
                            )

                            if re_enter_status:
                                try:
                                    await self._enter_option_trade(
                                        now,
                                        candle,
                                        trend,
                                        "Re-Entry",
                                        single_type=missing_leg,
                                        prefix="Re-entry paper trade"
                                    )
                                except Exception as e:
                                    self.logger.error(f"Error during {missing_leg} re-entry: {str(e)}")
                                    self.send_msg(f"Error during {missing_leg} re-entry: {str(e)}")
                                    traceback.print_exc()

                        elif self.both_sl_hit_in_trend:
                            self.logger.info(
                                f"Re-entry check after both SL hit: "
                                f"CE Condition {'met' if ce_re_enter else 'not met'}, "
                                f"PE Condition {'met' if pe_re_enter else 'not met'}, "
                                f"Entry Open={self.entry_open:.2f}, Entry Close={self.entry_close:.2f}, "
                                f"Current Close={close_price:.2f}"
                            )

                            if ce_re_enter or pe_re_enter:
                                try:
                                    if ce_re_enter:
                                        await self._enter_option_trade(
                                            now,
                                            candle,
                                            trend,
                                            "Re-Entry After Both SL",
                                            single_type="CE",
                                            prefix="Re-entry paper trade"
                                        )
                                    if pe_re_enter:
                                        await self._enter_option_trade(
                                            now,
                                            candle,
                                            trend,
                                            "Re-Entry After Both SL",
                                            single_type="PE",
                                            prefix="Re-entry paper trade"
                                        )
                                    self.both_sl_hit_in_trend = False
                                except Exception as e:
                                    self.logger.error(f"Error during re-entry after both SL: {str(e)}")
                                    self.send_msg(f"Error during re-entry after both SL: {str(e)}")
                                    traceback.print_exc()

                    if not self.active_trades and not self.time_exit_done:
                        if self.wait_for_new_candle:
                            next_candle_time = get_next_candle_close_time(
                                now, self.interval, self.intervals_config, self.market_config
                            )
                            sleep_seconds = (next_candle_time - now).total_seconds()
                            if sleep_seconds > 1:
                                self.logger.info(
                                    f"Waiting {sleep_seconds:.2f} sec for next candle close before entering new trades (restart logic)"
                                )
                                time.sleep(sleep_seconds)
                            self.wait_for_new_candle = False
                            continue

                        if (now - candle_time).total_seconds() < 0:
                            time.sleep(5)
                            continue

                        await self._enter_option_trade(
                            now, candle, trend, "New Trend", prefix="New paper trade"
                        )

                now_dt = datetime.now(ist)
                if self.last_pnl_report_time is None or (
                    now_dt - self.last_pnl_report_time
                ).total_seconds() >= 1200:
                    pnl_msg = self.calculate_running_pnl()
                    self.send_msg(pnl_msg)
                    self.last_pnl_report_time = now_dt

                next_candle_time = get_next_candle_close_time(
                    now, self.interval, self.intervals_config, self.market_config
                )
                sleep_now = datetime.now(ist)
                sleep_seconds = (next_candle_time - sleep_now).total_seconds()
                self.logger.info(
                    f"Sleeping for {sleep_seconds:.2f} seconds until next candle close at {next_candle_time}"
                )
                if sleep_seconds > 0:
                    time.sleep(sleep_seconds)
                else:
                    self.logger.info("Processing took longer than candle interval, skipping sleep.")


def main():
    ist = ZoneInfo(IST_ZONE)
    holidays = load_holidays()

    try:
        (
            interval,
            sma_period,
            telegram_bot_token,
            telegram_chat_id,
            symbol_config,
            paper_trading,
            paper_lot,
            market_config,
            intervals_config,
            matched_csv_path
        ) = load_configurations()

        weekly_off_days = market_config.get("weekly_off_days", ["SAT", "SUN"])
        today = datetime.now(ist).date()

        if not is_trading_day(today, holidays, weekly_off_days):
            logger.info("Today is not a trading day. Exiting.")
            return

        logger.info(f"ULTAPULTA LOGIN {today}")

    except Exception as e:
        logger.error(f"Failed to load configurations: {e}\n{traceback.format_exc()}")
        return

    paper_trader = PaperTrader(
        interval=interval,
        sma_period=sma_period,
        symbol_config=symbol_config,
        telegram_bot_token=telegram_bot_token,
        telegram_chat_id=telegram_chat_id,
        paper_trading=paper_trading,
        paper_lot=paper_lot,
        market_config=market_config,
        intervals_config=intervals_config,
        matched_csv_path=matched_csv_path
    )

    tick_sub = TickSubscriber(paper_trader)
    tick_sub.start()

    asyncio.run(
        paper_trader.run(
            ist=ZoneInfo(IST_ZONE),
            holidays=holidays,
            today=today,
            sl_monitor=None
        )
    )


if __name__ == "__main__":
    main()